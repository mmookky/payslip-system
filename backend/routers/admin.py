from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from auth import create_access_token, verify_password, get_current_admin
import models
import schemas
import openpyxl
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
from dateutil import parser as date_parser

router = APIRouter(prefix="/admin", tags=["Admin"])

COLUMN_MAP = {
    "No":                                           "no",
    "เลขรหัส / ID":                                "employee_code",
    "ชื่อ-สกุล / Name":                            "full_name",
    "ID Card":                                      "national_id",
    "แผนก / Department":                           "department",
    "ตำแหน่ง / Position":                          "position",
    "วันเริ่มทำงาน / Start Working":               "start_date",
    "เงินเดือน / Salary":                          "base_salary",
    "Paid Salary":                                  "paid_salary",
    "Allowance":                                    "allowance",
    "จำนวนชั่วโมงทำงานล่วงเวลา / Total Overtime": "overtime_hours",
    "สวัสดิการรวมทั้งหมด / Total Allowance":      "total_allowance",
    "รายการปรับ / Adjust":                         "adjust",
    "Special":                                      "special",
    "รายรับทั้งหมด / Total Income":                "total_income",
    "รายการหัก /Deduct":                           "deduct",
    "ประกันสังคม":                                  "social_security",
    "TAX":                                          "tax",
    "กองทุนสำรองเลี้ยงชีพ":                       "provident_fund",
    "หักลดหย่อน":                                  "tax_allowance",
    "กองทุนสงเคราะห์ลูกจ้าง":                     "welfare_fund",
    "รวมรายการหัก":                                "total_deductions",
    "รวมทั้งหมด / Total Payment":                  "net_pay",
    "รายได้สะสมประจำปี 2568":                      "ytd_income",
    "ประกันสังคมสะสม ประจำปี 2568":               "ytd_sso",
    "Total TAX":                                    "ytd_tax",
}

MANDATORY_FIELDS = [
    "employee_code",
    "full_name",
    "national_id",
    "department",
    "position",
    "base_salary",
    "paid_salary",
    "net_pay",
    # "ytd_tax",
]


def parse_number(val):
    if val is None or val == "":
        return None
    if isinstance(val, float) and val == int(val):
        return int(val)
    return val


def parse_date(val):
    if not val:
        return None
    try:
        if isinstance(val, datetime):
            return val.date()
        return date_parser.parse(str(val)).date()
    except Exception:
        return None


def parse_national_id(val):
    if not val:
        return None
    try:
        return str(int(float(str(val)))).strip()
    except Exception:
        return str(val).strip()


def validate_row(field_map: dict) -> list[str]:
    errors = []

    # Check mandatory fields
    for field in MANDATORY_FIELDS:
        val = field_map.get(field)
        if val is None or str(val).strip() == "" or str(val) == "None":
            col_name = next((k for k, v in COLUMN_MAP.items() if v == field), field)
            errors.append(f"'{col_name}' is required")

    # Check national_id must be numeric only
    national_id = field_map.get("national_id")
    if national_id:
        nid = parse_national_id(national_id)
        if not nid or not nid.isdigit():
            errors.append("'ID Card' must be numeric only")
        elif len(nid) != 13:
            errors.append(f"'ID Card' must be 13 digits (found {len(nid)} digits)")

    # Check salary fields must be numeric
    for field, label in [
        ("base_salary", "Salary"),
        ("paid_salary", "Paid Salary"),
        ("net_pay", "Total Payment")
    ]:
        val = field_map.get(field)
        if val is not None:
            try:
                float(str(val))
            except (ValueError, TypeError):
                errors.append(f"'{label}' must be numeric")

    return errors


@router.post("/login")
def admin_login(data: schemas.AdminLogin, db: Session = Depends(get_db)):
    admin = db.query(models.Admin).filter(models.Admin.username == data.username).first()
    if not admin or not verify_password(data.password, admin.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid username or password"
        )
    token = create_access_token({"sub": admin.username, "role": "admin"})
    return {"access_token": token, "token_type": "bearer", "role": "admin"}


@router.post("/upload", response_model=schemas.UploadResult)
def upload_payslip(
    month: int,
    year: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):
    # Check duplicate upload for completed month
    existing = db.query(models.UploadHistory).filter_by(
        month=month, year=year, status="completed"
    ).first()
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"Data for {month}/{year} already exists. Cannot upload again."
        )

    # Validate file type
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Only .xlsx and .xls are allowed."
        )

    contents = file.file.read()

    try:
        wb = openpyxl.load_workbook(BytesIO(contents), read_only=True)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Cannot read Excel file. Please check the file format."
        )

    ws = wb.active

    # Read header row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers.append(str(cell).strip() if cell else "")

    # Map header → index
    col_index = {}
    unknown_cols = []
    for i, h in enumerate(headers):
        if h in COLUMN_MAP:
            col_index[COLUMN_MAP[h]] = i
        elif h and h != "None":
            unknown_cols.append(h)

    global_errors = []
    if unknown_cols:
        global_errors.append(f"Unknown columns found (will be skipped): {', '.join(unknown_cols)}")

    # Read and validate all rows
    all_rows = []
    row_errors = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        def get_raw(field, r=row):
            idx = col_index.get(field)
            val = r[idx] if idx is not None else None
            if val == "" or val is None:
                return None
            return parse_number(val)

        field_map = {field: get_raw(field) for field in COLUMN_MAP.values()}
        errors = validate_row(field_map)

        all_rows.append((row_idx, row, field_map))
        if errors:
            row_errors[row_idx] = errors

    if len(all_rows) == 0:
        raise HTTPException(
            status_code=400,
            detail="No data found in the Excel file."
        )

    # Check duplicates within the file
    seen_codes = {}
    seen_nationals = {}
    for row_idx, row, field_map in all_rows:
        code = str(parse_national_id(field_map.get("employee_code")) or "")
        nid = parse_national_id(field_map.get("national_id"))

        if code and code in seen_codes:
            existing_errors = row_errors.get(row_idx, [])
            existing_errors.append(f"'เลขรหัส / ID' value '{code}' is duplicated in this file (first seen at row {seen_codes[code]})")
            row_errors[row_idx] = existing_errors
        elif code:
            seen_codes[code] = row_idx

        if nid and nid in seen_nationals:
            existing_errors = row_errors.get(row_idx, [])
            existing_errors.append(f"'ID Card' value '{nid}' is duplicated in this file (first seen at row {seen_nationals[nid]})")
            row_errors[row_idx] = existing_errors
        elif nid:
            seen_nationals[nid] = row_idx

    total = len(all_rows)
    failed = len(row_errors)

    # If any row fails → do not save, return error file
    if failed > 0:
        error_wb = Workbook()
        error_ws = error_wb.active
        error_ws.title = "Upload Errors"
        error_ws.append(headers + ["Error"])

        for row_idx, row, field_map in all_rows:
            errors = row_errors.get(row_idx, [])
            error_msg = " | ".join(errors) if errors else ""
            row_list = list(row)
            # Fix national_id display in error file
            nid_idx = col_index.get("national_id")
            if nid_idx is not None and row_list[nid_idx]:
                row_list[nid_idx] = parse_national_id(row_list[nid_idx])
            error_ws.append(row_list + [error_msg])

        error_buffer = BytesIO()
        error_wb.save(error_buffer)
        error_bytes = error_buffer.getvalue()

        upload = models.UploadHistory(
            admin_id=current_admin.id,
            filename=file.filename,
            original_file=contents,
            error_file=error_bytes,
            month=month,
            year=year,
            total_records=total,
            success_records=0,
            failed_records=failed,
            status="failed"
        )
        db.add(upload)
        db.commit()
        db.refresh(upload)

        return schemas.UploadResult(
            id=upload.id,
            filename=file.filename,
            month=month,
            year=year,
            total_records=total,
            success_records=0,
            failed_records=failed,
            status="failed",
            has_error_file=True,
            errors=global_errors + [
                f"Row {row_idx}: {' | '.join(errs)}"
                for row_idx, errs in row_errors.items()
            ]
        )

    # All rows valid → save all
    upload = models.UploadHistory(
        admin_id=current_admin.id,
        filename=file.filename,
        original_file=contents,
        error_file=None,
        month=month,
        year=year,
        status="processing"
    )
    db.add(upload)
    db.flush()

    for row_idx, row, field_map in all_rows:
        def get(field, fm=field_map):
            return fm.get(field)

        national_id = parse_national_id(get("national_id"))

        employee = db.query(models.Employee).filter_by(national_id=national_id).first()
        if not employee:
            full_name = str(get("full_name") or "").strip()
            parts = full_name.split(" ", 1)
            first_name = parts[0] if parts else full_name
            last_name = parts[1] if len(parts) > 1 else ""

            from auth import hash_password
            employee = models.Employee(
                national_id=national_id,
                employee_code=str(parse_national_id(get("employee_code")) or national_id),
                first_name=first_name,
                last_name=last_name,
                department=get("department"),
                position=get("position"),
                start_date=parse_date(get("start_date")),
                password_hash=hash_password(national_id)
            )
            db.add(employee)
            db.flush()

        payslip = models.Payslip(
            employee_id=employee.id,
            upload_id=upload.id,
            month=month,
            year=year,
            base_salary=get("base_salary") or 0,
            paid_salary=get("paid_salary") or 0,
            allowance=get("allowance") or 0,
            overtime_hours=get("overtime_hours") or 0,
            total_allowance=get("total_allowance") or 0,
            adjust=get("adjust") or 0,
            special=get("special") or 0,
            total_income=get("total_income") or 0,
            deduct=get("deduct") or 0,
            social_security=get("social_security") or 0,
            tax=get("tax") or 0,
            provident_fund=get("provident_fund") or 0,
            tax_allowance=get("tax_allowance") or 0,
            welfare_fund=get("welfare_fund") or 0,
            total_deductions=get("total_deductions") or 0,
            net_pay=get("net_pay") or 0,
            ytd_income=get("ytd_income") or 0,
            ytd_tax=get("ytd_tax") or 0,
            ytd_sso=get("ytd_sso") or 0,
        )
        db.add(payslip)

    upload.total_records = total
    upload.success_records = total
    upload.failed_records = 0
    upload.status = "completed"
    db.commit()
    db.refresh(upload)

    return schemas.UploadResult(
        id=upload.id,
        filename=file.filename,
        month=month,
        year=year,
        total_records=total,
        success_records=total,
        failed_records=0,
        status="completed",
        has_error_file=False,
        errors=global_errors
    )


@router.get("/uploads", response_model=list[schemas.UploadHistoryOut])
def get_upload_history(
    db: Session = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):
    return db.query(models.UploadHistory).order_by(
        models.UploadHistory.uploaded_at.desc()
    ).all()


@router.get("/uploads/{upload_id}/download-original")
def download_original_file(
    upload_id: int,
    db: Session = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):
    upload = db.query(models.UploadHistory).filter_by(id=upload_id).first()
    if not upload or not upload.original_file:
        raise HTTPException(status_code=404, detail="File not found")
    return StreamingResponse(
        BytesIO(upload.original_file),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={upload.filename}"}
    )


@router.get("/uploads/{upload_id}/download-error")
def download_error_file(
    upload_id: int,
    db: Session = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):
    upload = db.query(models.UploadHistory).filter_by(id=upload_id).first()
    if not upload or not upload.error_file:
        raise HTTPException(status_code=404, detail="Error file not found")
    error_filename = f"error_{upload.filename}"
    return StreamingResponse(
        BytesIO(upload.error_file),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={error_filename}"}
    )


@router.get("/payslips", response_model=list[schemas.AdminPayslipOut])
def get_payslips_by_month(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    current_admin: models.Admin = Depends(get_current_admin)
):
    return db.query(models.Payslip).filter_by(
        month=month, year=year
    ).all()